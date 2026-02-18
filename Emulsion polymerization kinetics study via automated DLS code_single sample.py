from datetime import datetime
import serial
import serial.tools.list_ports
from time import *
from time import sleep
import os
from SF10 import SF10
from syringepump import SyringePump
from switchValve import *
import subprocess
import pyautogui
import pygetwindow as gw 
import glob
import os.path
import pandas as pd
import numpy as np
from pandas import ExcelWriter
from pandas import ExcelFile
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.ticker as ticker
from matplotlib import cm
import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from shutil import copyfile
from colorama import Fore, Style
from collections import Counter

# Find the COMs that syringe pumps connected to computer
PortData = serial.tools.list_ports.comports()
print(PortData)
for port in PortData:
    print(f"\033[1;31m{port}\033[0m")
    
# Pumps Assigned 
PumpFeed = SF10('COM9', 'PumpFeed')
PumpQuench = SyringePump('COM4', 'PumpQuench')
PumpDilution = SyringePump('COM7','PumpDilution')
switchvalve = SwitchValve('COM5','Switchvalve')

#entering operator name and solvents
operator = input('Please input the name of the operator:>> ')

# Get the current date and time
current_datetime = datetime.now()
# Format the datetime as a string
time_for_report = current_datetime.strftime("%Y-%m-%d %H-%M-%S")

#saving the input file in reports folder
def save_inputexcel(original_path, new_folder, new_name):
    df_input = pd.read_excel(original_path, sheet_name='Sheet1')  
    new_path = os.path.join(new_folder, new_name)
    df_input.to_excel(new_path, index=False)
    print(Fore.GREEN +  f"Input Excel sheet copied and saved to: {new_path}")
    print(Style.RESET_ALL)
    return new_path

original_excel_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\flow parameters setup.xlsx"
new_folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Reports"
new_excel_name = "Input Report_{}_{}.xlsx".format(operator, time_for_report) 
saved_input_path = save_inputexcel(original_excel_path, new_folder_path, new_excel_name)

#creating an output report for the experiment
path_to_report = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Reports"
report_workbook = Workbook()
sheet = report_workbook.active  # Get the active sheet (default sheet)
# Add data to the sheet
sheet["A1"] = "Date :"
sheet["A1"].font = Font(bold=True) #making the text bold
sheet["B1"] = datetime.now().date() # adding the current date
sheet["D1"] = "Operator :"
sheet["D1"].font = Font(bold=True) #making the text bold
sheet["E1"] = operator

sheet["A3"] = "Input :"
sheet["A3"].font = Font(bold=True)
sheet["P3"] = "Output :"
sheet["P3"].font = Font(bold=True)

output_titles = {"A4" : "Experiment no.", "B4" : "polymer type", "C4" : "solids content (wt%)", "D4" : "[BRC] wt% bom", "E4" : "[APS] wt% bom", 
           "F4" : "Temperature (C)", "G4" : "base", "H4" : "additional [base] mM", "I4" : "reactor volume (mL)", "J4" : "residence time (min)","K4" : "repitition", 
           "L4" : "flow rate of feed pump (mL/min)", "M4" : "flow rate of quench pump (mL/min)", "N4" : "dilution pump flow rate (mL/min)",
           "P4" : "Measurement name", "Q4" : "Hydrodynamic diameter (µm)", "R4" : "Polydispersity index", 
           "S4" : "Baseline", "T4" : "Peak volume 1 (µm)", "U4" : "Area volume 1 %", "V4" : "Peak volume 2 (µm)", "W4" : "Area volume 2 %", 
           "X4" : "Peak volume 3 (µm)", "Y4" : "Area volume 3 %", "Z4" : "Peak intensity 1 (µm)", "AA4" : "Area intensity 1 %", "AB4" : "Peak intensity 2 (µm)", 
           "AC4" : "Area intensity 2 %", "AD4" : "Peak intensity 3 (µm)", "AE4" : "Area intensity 3 %", "AF4" : "Peak number 1 (µm)", "AG4" : "Area number 1 %", 
           "AH4" : "Peak number 2 (µm)", "AI4" : "Area number 2 %", "AJ4" : "Peak number 3 (µm)", "AK4" : "Area number 3 % "}
for cell, value in output_titles.items():
    sheet[cell] = value
    sheet[cell].font = Font(bold=True, color="000080")
    
report_name = "Output Report_{}_{}.xlsx".format(operator, time_for_report)  
report_path = os.path.join(path_to_report, report_name)
report_workbook.save(report_path)
print(Fore.GREEN +  f"Output Excel sheet created and saved to: {report_path}")
print(Style.RESET_ALL)

#adding input and output files to the report summary in reports folder
summary_path1 = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Reports\Report summary.xlsx"
summary_sheet = "Sheet1"
input_path = saved_input_path
output_path = report_path
summary_workbook = openpyxl.load_workbook(summary_path1)
summary_sheet = summary_workbook[summary_sheet]
summary_sheet.cell(row=summary_sheet.max_row + 1, column=1, value=current_datetime)
summary_sheet.cell(row=summary_sheet.max_row, column=2, value=operator)
summary_sheet.cell(row=summary_sheet.max_row, column=3, value=input_path)
summary_sheet.cell(row=summary_sheet.max_row, column=4, value=output_path)
summary_workbook.save(summary_path1)
print(Fore.GREEN + f"Input and Output Excel sheet links added and saved in: {summary_path1}")
print(Style.RESET_ALL)
   
# Flow rates, mixing profiles and repetitons input

#Reading excel data for input
EXP_NO = "Experiment no."
POLYMER_TYPE = "polymer type"
SOLIDS_CONTENT = "solids content (wt%)"
BRC_CONCENTRATION = "[BRC] wt% bom"
APS_CONCENTRATION = "[APS] wt% bom"
TEMPERATURE = "Temperature (C)"
BASE = "base"
BASE_CONCENTRATION = "additional [base] (mM)"
REACTOR_VOLUME = "reactor volume (mL)"
RESIDENCE_TIME = "residence time"
REP = "repetition"
FLOW_RATE_FEED = "flow rate of feed pump (mL/min)"
FLOW_RATE_QUENCH = "flow rate of quench pump (mL/min)"
FLOW_RATE_DILUTION = "dilution pump flow rate (mL/min)"

df_excel = pd.read_excel(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\flow parameters setup.xlsx")
#print(df_excel)
numeric_columns = df_excel.select_dtypes(include=[float])
df_excel[numeric_columns.columns] = df_excel[numeric_columns.columns].round(2)
row_names = df_excel.index
row_names_list = df_excel.index.tolist()

def flow_parameters(a, b, c, d, e, f, g, h):

    #making a new list including repetitions required for feed pump rates   
    newlist_feedpump = [ratef for item, ratef in zip(c, a) for repnumber in range(item)]
                    
    #making a new list including repetitions required for quench pump and dilution pump flow rates
    newlist_quenchpump = [rateq for item, rateq in zip(c, b) for repnumber in range(item)]
    newlist_dilutionpump = [rated for item, rated in zip(c, d) for repnumber in range(item)] 
    
    #making a new list including repetitions required for indexes 
    newlist_index = [ratei for item, ratei in zip(c, e) for repnumber in range(item)]
    
    #making a new list including repetitions required for polymer type
    newlist_polymername = [raten for item, raten in zip(c, f) for repnumber in range(item)]

    #making a new list including repetitions required for reactor volume
    newlist_reactorvolume = [ratev for item, ratev in zip(c, g) for repnumber in range(item)]
    
    #making a new list including repetitions required for experiment number
    newlist_expno = [ratee for item, ratee in zip(c, h) for repnumber in range(item)]
 
    for flowrates in zip(newlist_feedpump,newlist_quenchpump, newlist_dilutionpump,newlist_index,newlist_polymername,newlist_reactorvolume, newlist_expno):
        print (flowrates)  #combines five lists for five iterations at the same time
    
    #To perform sample stabilization only in the first run    
    previous_exp_no = None
    
    for flowrates,(fp,qp,dp,pi,pn,rv,en) in enumerate (zip(newlist_feedpump,newlist_quenchpump, newlist_dilutionpump,newlist_index,newlist_polymername,newlist_reactorvolume, newlist_expno)):
        print (flowrates,fp,qp,dp,pi,pn,rv,en) #gives an index to the zipped output as in a list
    
        print(Fore.GREEN + "Experiment {} is starting : flow rate of feed pump : {} ml/min, quench pump : {} ml/min, dilution pump : {} ml/min for {} at {} ml reactor volume " .format(flowrates+1,fp,qp,dp,pn,rv))
        print(Style.RESET_ALL)
        
        #printing the input row in the output workbook final row
        load_workbook(report_path) #load the workbook
        sheet = report_workbook.active  
        target_row_number = sheet.max_row + 1
        source_row_index = pi  
        row_to_print = df_excel.iloc[source_row_index]
         
        for col_idx, cell_value in enumerate(row_to_print):
            sheet.cell(row=target_row_number, column=col_idx + 1, value=cell_value)
        report_workbook.save(report_path)

        # Experiment condition input  
        V_total = 1.75
        #sleeptime = time for which the pumps run flling the cell and tubing with micelles
        sleeptime = (V_total*60)/fp
        
        switchvalve.toPositionA()
        
        print(Fore.GREEN + "Dilution pump will start at {} ml/min flow rate".format(dp))
        print(Style.RESET_ALL)

        # starting the dlilution pump for flushing
        PumpDilution.start()
        sleep(0.5)
        PumpDilution.changeFlowrate(dp)
        sleep(0.5)
        PumpDilution.start()
        sleep(60)
        PumpDilution.stop()
            
        print(Fore.GREEN + "This is the first run. Feed pump will start at {} ml/min and quench pump will start at {} ml/min flow rate".format(fp,qp))
        print(Style.RESET_ALL)
            
        # starting the feed pump and quench pump
        PumpFeed.start()
        sleep(0.5)
        PumpFeed.changeFlowrate(fp) 
            
        PumpQuench.start()
        sleep(0.5)
        PumpQuench.changeFlowrate(qp)
        sleep(0.5)
        PumpQuench.start()
        sleep(0.5)
        
        #switch valve to position A for 1.33*residence time
        switchvalve.toPositionA()
        print(f"Stabilizing for the first run of experiment {en}")
        sleep(1.4*sleeptime)

 #switch valve changes to position B for directing sample to DLS
        switchvalve.toPositionB()
        
        # starting the dlilution pump
        PumpDilution.start()
        sleep(0.5)
        PumpDilution.changeFlowrate(dp)
        sleep(0.5)
        PumpDilution.start()
        sleep(15)
        
        print(Fore.GREEN + "Switch valve changing to pos A and dilution pump will stop")
        print(Style.RESET_ALL)
         
        #switch valve to position A for DLS analysis
        switchvalve.toPositionA()
        PumpDilution.stop()
        
        print(Fore.GREEN + "Experiment {} is finished. DLS analysis will start now".format(flowrates+1))
        print(Style.RESET_ALL)
        #DLS analysis 

        #activate kalliope screen
        sleep(3)
        #x=gw.getAllTitles()
        #print(x)
        hwnd = gw.getWindowsWithTitle('apkw')
        print(hwnd)
        if hwnd != []:
            try:
                hwnd[0].activate()
            except:
                hwnd[0].minimize()
                hwnd[0].maximize()

        #click on copy prameters button
        sleep(4)
        copy_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\copy1.png", confidence=0.5) 
        print (copy_click)
        pyautogui.moveTo(copy_click,duration=2)
        pyautogui.click(copy_click)

        #erases the current title
        sleep(2)
        pyautogui.hotkey("backspace")

        #types the title including current exp no, polymer name, water flow rate and polymer flow rate
        pyautogui.write("Exp {}_{}_feed {}_quench {}_dilution {}_RV {}".format(flowrates+1,pn,fp,qp,dp,rv))

        #click on start button
        sleep(3)
        start_click = pyautogui.locateCenterOnScreen("S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\start1.png", confidence=0.5) 
        print (start_click)
        pyautogui.moveTo(start_click,duration=2)
        pyautogui.click(start_click)

        #Monitoring for a new file generation in excel files folder
        path_to_watch = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        print(Fore.GREEN + "The path of the folder where your excel data will be saved is", path_to_watch)
        print(Style.RESET_ALL)
        before = dict ([(f, None) for f in os.listdir (path_to_watch)])
        while 1:
            after = dict ([(f, None) for f in os.listdir (path_to_watch)])
            added = [f for f in after if not f in before]
            if added:
                    print("Added: ", ", ".join (added))
                    break
            else:
                    before = after

        #identifying the most recent excel file in the folder
        folder_path = r"S:\Sci-Chem\PRD\DLS\DLS-Data\Excel files"
        file_type = r'\*xlsx'
        files = glob.glob(folder_path + file_type)
        max_file = max(files, key=os.path.getctime)

        print(Fore.GREEN + "The path of the folder where the size distribution plots are saved is", r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Plots")
        print(Style.RESET_ALL)

        #creating the dataframe for plotting
        sleep(2)
        df = pd.read_excel(max_file)
        #print(df)
        
        df1=df[["Unnamed: 5","Unnamed: 6","Unnamed: 7","Unnamed: 8"]]
        df1.loc[:,'Unnamed: 5']*=1000
        df1.columns=["Particle diameter","Intensity weighted","Volume weighted","Number weighted"]
        df1=df1.dropna(axis=0,how="any")
        df1.drop([4,6],axis=0,inplace=True)
        df1=df1.astype(float)
        df1["Particle diameter"] = df1["Particle diameter"].map('{:.1f}'.format)
        #print(df1)
        time_for_plots = current_datetime.strftime("%Y-%m-%d %H-%M-%S")
        
        #Intensity Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Intensity weighted")
        plt.ylabel("Intensity Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Intensity Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Plots")  
        my_file = 'Exp {}_{}_Intensity weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #Volume Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Volume weighted")
        plt.ylabel("Volume Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Volume Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Plots") 
        my_file = 'Exp {}_{}_Volume weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #Number Weighted size distribution plot
        df1.plot.bar(x="Particle diameter",y="Number weighted")
        plt.ylabel("Number Weighted %",fontsize=12)
        plt.xlabel("Particle Size (nm)",fontsize=12)
        plt.title("Number Weighted size distribution")
        plt.gca().set_xticks(plt.gca().get_xticks()[::4])
        figure = plt.gcf() # get current figure
        figure.set_size_inches(6,9)
        my_path = os.path.abspath(r"S:\Sci-Chem\PRD\DLS\DLS-Data\Sandra\Plots") 
        my_file = 'Exp {}_{}_Number weighted_{}.png'.format(flowrates+1,pn,time_for_plots)
        plt.savefig(os.path.join(my_path, my_file))  
        plt.close()
        
        #exporting specific data from exported file to output report excel file
        source_file_path = max_file  # the path to source Excel file
        #df_source = pd.read_excel(max_file)
        source_workbook = openpyxl.load_workbook(source_file_path)
        source_sheet = source_workbook['Measurement 0']

        data_series = [
            source_sheet['B2'].value,
            source_sheet['C7'].value,
            source_sheet['C8'].value,
            source_sheet['C10'].value,
            source_sheet['C15'].value,
            source_sheet['C16'].value,
            source_sheet['C18'].value,
            source_sheet['C19'].value,
            source_sheet['C21'].value,
            source_sheet['C22'].value,
            source_sheet['C24'].value,
            source_sheet['C25'].value,
            source_sheet['C27'].value,
            source_sheet['C28'].value,
            source_sheet['C30'].value,
            source_sheet['C31'].value,
            source_sheet['C33'].value,
            source_sheet['C34'].value,
            source_sheet['C36'].value,
            source_sheet['C37'].value,
            source_sheet['C39'].value,
            source_sheet['C40'].value  
        ]

        openpyxl.load_workbook(report_path)
        sheet = report_workbook.active
        columns_to_append = ['P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK']

        for cellvalue, column in enumerate(columns_to_append):
            cell_address = f"{column}{sheet.max_row}"
            sheet[cell_address] = data_series[cellvalue]

        report_workbook.save(report_path)
        print(Fore.GREEN + f"The data is extracted and saved in {report_path}")
        print(Style.RESET_ALL)

flow_parameters(list(df_excel[FLOW_RATE_FEED]), list(df_excel[FLOW_RATE_QUENCH]), list(df_excel[REP]), list(df_excel[FLOW_RATE_DILUTION]), row_names_list, list(df_excel[POLYMER_TYPE]), list(df_excel[REACTOR_VOLUME]), list(df_excel[EXP_NO]))

PumpFeed.stop(), PumpQuench.stop()

print(Fore.GREEN + "Main experiment is over. Please check 1_excel files and 2_reports folder and 3_plots folders in Sandra's folder for data")
print(Style.RESET_ALL)        